import asyncio
from starlette.applications import Starlette
from starlette.routing import WebSocketRoute
from starlette.websockets import WebSocketDisconnect
from starlette.websockets import WebSocket
import time
import os
import cv2
from datetime import datetime
from ultralytics import YOLO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.image import Image
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import requests
from PIL import Image as PILImage
from threading import Thread

print("initialize...")

# Inisialisasi model YOLO dan sumber video
model = YOLO("trained_models/v5-new_64_50.pt")
cap = cv2.VideoCapture("rtsp://10.3.51.251/live/ch00_0")
wb = Workbook()
ws = wb.active
col_time = "A"
col_type = "B"
col_image = "C"

# Konfigurasi bot Telegram
TOKEN = "7917230413:AAEYJx9PnSTo9IyYvezMSiEdAsYB9kak_-A"
CHAT_ID = "-1002437276925"
status_detection = True  # Status deteksi awal

# Waktu mulai program
start_time = datetime.now().strftime("%H-%M-%S %d-%m-%Y")


def send_telegram_message(message):
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    data = {"chat_id": CHAT_ID, "text": message}
    try:
        response = requests.post(url, data=data)
        if response.status_code == 200:
            print("Peringatan terkirim ke Telegram.")
        else:
            print(f"Gagal mengirim pesan. Status code: {response.status_code}")
    except Exception as e:
        print(f"Error saat mengirim pesan: {e}")


def send_telegram_photo(photo_path):
    url = f"https://api.telegram.org/bot{TOKEN}/sendPhoto"
    files = {"photo": open(photo_path, "rb")}
    data = {"chat_id": CHAT_ID}
    try:
        response = requests.post(url, files=files, data=data)
        if response.status_code == 200:
            print("Foto terkirim ke Telegram.")
        else:
            print(f"Gagal mengirim foto. Status code: {response.status_code}")
    except Exception as e:
        print(f"Error saat mengirim foto: {e}")


def listen_to_bot():
    global status_detection
    url = f"https://api.telegram.org/bot{TOKEN}/getUpdates"
    last_update_id = None

    while True:
        try:
            response = requests.get(url)
            if response.status_code == 200:
                updates = response.json().get("result", [])
                if updates:
                    for update in updates:
                        update_id = update["update_id"]
                        message = update.get("message", {}).get("text", "").lower()

                        if last_update_id is None or update_id > last_update_id:
                            last_update_id = update_id
                            if message == "matikan":
                                status_detection = False
                                send_telegram_message("Sistem deteksi telah dimatikan.")
                            elif message == "nyalakan":
                                status_detection = True
                                send_telegram_message("Sistem deteksi telah dihidupkan.")
        except Exception as e:
            print(f"Error saat mendengarkan bot: {e}")
        time.sleep(2)


def process(frame, type, start_time, current_time, row):
    folder_path = f"pri_images/{type}"
    os.makedirs(folder_path, exist_ok=True)

    pict = f"{folder_path}/{current_time}.jpg"
    cv2.imwrite(pict, frame)

    img = Image(pict)
    img.height = 300
    img.width = 400
    ws[f"{col_time}{row}"] = current_time
    ws[f"{col_type}{row}"] = type
    ws.add_image(img, f"{col_image}{row}")
    wb.save(f"result_{start_time}.xlsx")
    generate_pdf(type, current_time, pict)
    Thread(target=send_telegram_message, args=(f"Terdeteksi kejadian {type.capitalize()} pada {current_time}. Mohon segera diperiksa kondisi terkininya",)).start()
    Thread(target=send_telegram_photo, args=(pict,)).start()


def generate_pdf(type, current_time, image_path):
    pdf_filename = f"report_{type}_{current_time}.pdf"
    c = canvas.Canvas(pdf_filename, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, height - 50, "LAPORAN DETEKSI FALL")

    c.setFont("Helvetica", 12)
    c.drawString(100, height - 100, f"Waktu dan Tanggal Kejadian: {current_time}")
    c.drawString(100, height - 120, "Tempat Kejadian: Tangga Teknik Fisika ITS")
    c.drawString(100, height - 140, f"Jenis Kejadian: {type.capitalize()}")

    c.setFont("Helvetica-Bold", 14)
    c.drawString(100, height - 170, "Dokumentasi Kejadian:")

    try:
        image = PILImage.open(image_path)
        image = image.resize((400, 300))
        temp_image_path = "resized_image.jpg"
        image.save(temp_image_path)
        c.drawImage(temp_image_path, 100, height - 500)
        os.remove(temp_image_path)
    except Exception as e:
        print(f"Error saat memuat gambar: {e}")

    c.setFont("Helvetica-Oblique", 10)
    c.setFillColor(colors.red)
    c.drawString(
        100, 100, "Terdeteksi Kejadan. Mohon segera diperiksa kondisi terkininya."
    )

    c.save()


Thread(target=listen_to_bot, daemon=True).start()

async def app(scope, receive, send):
    websocket = WebSocket(scope=scope, receive=receive, send=send)
    await websocket.accept()
    await asyncio.sleep(5)
    try:
        row = 1
        prev_frame_time = time.time()
        fall_frame = 0
        frame_count = 0
        while cap.isOpened():
            now = datetime.now()
            current_time = now.strftime("%H-%M-%S %d-%m-%Y")
            ret, frame = cap.read()

            if not ret:
                print("Video frame is empty or video processing has been successfully completed.")
                break

            # Hitung FPS
            new_frame_time = time.time()
            fps = 1 / (new_frame_time - prev_frame_time)
            prev_frame_time = new_frame_time

            # Tampilkan status deteksi pada frame
            status_text = "Deteksi: ON" if status_detection else "Deteksi: OFF"
            cv2.putText(frame, status_text, (20, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 1)
            cv2.putText(frame, f"FPS: {int(fps)}", (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 1)

            if status_detection:
                results = model(frame, conf=0.8)

                for r in results:
                    frame = r.plot()
                    frame_count += 1
                    print("frame_count: ", frame_count)
                    class_id_list = list(r.boxes.cls)
                    if frame_count > 0 and 0 in class_id_list:
                        fall_frame += 1
                        print("fall_frame: ", fall_frame)
                        if fall_frame == 10:
                            print("=======================")
                            print("=====Fall detected=====")
                            print("=======================")
                            process(frame, 'fall', start_time, current_time)
                            await websocket.send_text("fall")
                            row += 20
                            fall_frame = 0
                            frame_count = -100
                    else:
                        fall_frame = 0
                            
                        

           
            cv2.imshow("frame", frame)
            if cv2.waitKey(1) == ord("q"):
                break

    except WebSocketDisconnect:
        print("WebSocket connection was disconnected.")
    except Exception as e:
        print(f"Error during detection: {e}")
    finally:
        cap.release()
        cv2.destroyAllWindows()
        await websocket.close()


# routes = [WebSocketRoute("/pri", endpoint=detect)]


