<<<<<<< HEAD
import os
import sys

=======
import asyncio
from starlette.applications import Starlette
from starlette.routing import WebSocketRoute
from starlette.websockets import WebSocketDisconnect
from starlette.websockets import WebSocket
import time
import os
import cv2
from datetime import datetime
from ultralytics import YOLO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.image import Image
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import requests
from PIL import Image as PILImage
from threading import Thread

import os
import sys
>>>>>>> 6708c07 (integrate w software)
import cv2
import joblib
import numpy as np
import pandas as pd
import tensorflow as tf
from keras._tf_keras.keras.models import load_model
from keras._tf_keras.keras.utils import pad_sequences
from ultralytics import YOLO
<<<<<<< HEAD
=======
from openpyxl import Workbook
>>>>>>> 6708c07 (integrate w software)

sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/../..")
from utils.coco_keypoints import COCOKeypoints, extract_keypoints

gpus = tf.config.experimental.list_physical_devices("GPU")
try:
    for gpu in gpus:
        tf.config.experimental.set_memory_growth(gpu, True)
except RuntimeError as e:
    print(e)

# Global variables
<<<<<<< HEAD
=======
cap = cv2.VideoCapture("rtsp://10.3.51.251/live/ch00_0")
>>>>>>> 6708c07 (integrate w software)
MIN_FRAMES = 10
MAX_TIMESTEPS = 30
rolling_buffer = {}

# Load models and scalers
<<<<<<< HEAD
model_path = "ml/lstm/slip_fall_detector.keras"
scaler_path = "ml/lstm/lstm_scaler.pkl"
lstm_model = load_model(model_path)
scaler = joblib.load(scaler_path)
yolo_model = YOLO("yolo11n-pose.engine")
=======
KEYPOINT_MODEL_PATH = "ml/yolo11n-pose.engine"
LSTM_MODEL_PATH = "ml/lstm/slip_fall_detector.keras"
FALL_MODEL_PATH = "ml/fall.engine"
SCALER_PATH = "ml/lstm/lstm_scaler.pkl"

lstm_model = load_model(LSTM_MODEL_PATH)
scaler = joblib.load(SCALER_PATH)
keypoint_model = YOLO(KEYPOINT_MODEL_PATH, task="pose")
fall_model = YOLO(FALL_MODEL_PATH, task="detect")

wb = Workbook()
ws = wb.active
col_time = "A"
col_type = "B"
col_image = "C"

# Konfigurasi bot Telegram
TOKEN = "7917230413:AAEYJx9PnSTo9IyYvezMSiEdAsYB9kak_-A"
CHAT_ID = "-1002437276925"
status_detection = True  # Status deteksi awal

# Waktu mulai program
start_time = datetime.now().strftime("%H-%M-%S %d-%m-%Y")
>>>>>>> 6708c07 (integrate w software)

base_folder = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "../../dataset_primer/"
)
target_size = (640, 480)
labels = ["Slip", "SlipFall", "Trip", "TripFall", "Walking"]

# Define columns without confidence values
COCO_KEYPOINTS_NO_CONF = [
    "nose_x",
    "nose_y",
    "left_eye_x",
    "left_eye_y",
    "right_eye_x",
    "right_eye_y",
    "left_ear_x",
    "left_ear_y",
    "right_ear_x",
    "right_ear_y",
    "left_shoulder_x",
    "left_shoulder_y",
    "right_shoulder_x",
    "right_shoulder_y",
    "left_elbow_x",
    "left_elbow_y",
    "right_elbow_x",
    "right_elbow_y",
    "left_wrist_x",
    "left_wrist_y",
    "right_wrist_x",
    "right_wrist_y",
    "left_hip_x",
    "left_hip_y",
    "right_hip_x",
    "right_hip_y",
    "left_knee_x",
    "left_knee_y",
    "right_knee_x",
    "right_knee_y",
    "left_ankle_x",
    "left_ankle_y",
    "right_ankle_x",
    "right_ankle_y",
]

<<<<<<< HEAD
=======
def send_telegram_message(message):
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    data = {"chat_id": CHAT_ID, "text": message}
    try:
        response = requests.post(url, data=data)
        if response.status_code == 200:
            print("Peringatan terkirim ke Telegram.")
        else:
            print(f"Gagal mengirim pesan. Status code: {response.status_code}")
    except Exception as e:
        print(f"Error saat mengirim pesan: {e}")


def send_telegram_photo(photo_path):
    url = f"https://api.telegram.org/bot{TOKEN}/sendPhoto"
    files = {"photo": open(photo_path, "rb")}
    data = {"chat_id": CHAT_ID}
    try:
        response = requests.post(url, files=files, data=data)
        if response.status_code == 200:
            print("Foto terkirim ke Telegram.")
        else:
            print(f"Gagal mengirim foto. Status code: {response.status_code}")
    except Exception as e:
        print(f"Error saat mengirim foto: {e}")


def listen_to_bot():
    global status_detection
    url = f"https://api.telegram.org/bot{TOKEN}/getUpdates"
    last_update_id = None

    while True:
        try:
            response = requests.get(url)
            if response.status_code == 200:
                updates = response.json().get("result", [])
                if updates:
                    for update in updates:
                        update_id = update["update_id"]
                        message = update.get("message", {}).get("text", "").lower()

                        if last_update_id is None or update_id > last_update_id:
                            last_update_id = update_id
                            if message == "matikan":
                                status_detection = False
                                send_telegram_message("Sistem deteksi telah dimatikan.")
                            elif message == "nyalakan":
                                status_detection = True
                                send_telegram_message("Sistem deteksi telah dihidupkan.")
        except Exception as e:
            print(f"Error saat mendengarkan bot: {e}")
        time.sleep(2)

def generate_pdf(type, current_time, image_path):
    pdf_filename = f"report_{type}_{current_time}.pdf"
    c = canvas.Canvas(pdf_filename, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, height - 50, "LAPORAN DETEKSI FALL")

    c.setFont("Helvetica", 12)
    c.drawString(100, height - 100, f"Waktu dan Tanggal Kejadian: {current_time}")
    c.drawString(100, height - 120, "Tempat Kejadian: Tangga Teknik Fisika ITS")
    c.drawString(100, height - 140, f"Jenis Kejadian: {type.capitalize()}")

    c.setFont("Helvetica-Bold", 14)
    c.drawString(100, height - 170, "Dokumentasi Kejadian:")

    try:
        image = PILImage.open(image_path)
        image = image.resize((400, 300))
        temp_image_path = "resized_image.jpg"
        image.save(temp_image_path)
        c.drawImage(temp_image_path, 100, height - 500)
        os.remove(temp_image_path)
    except Exception as e:
        print(f"Error saat memuat gambar: {e}")

    c.setFont("Helvetica-Oblique", 10)
    c.setFillColor(colors.red)
    c.drawString(
        100, 100, "Terdeteksi Kejadan. Mohon segera diperiksa kondisi terkininya."
    )

    c.save()

def process_telegram(frame, type, start_time, current_time, row):
    folder_path = f"pri_images/{type}"
    os.makedirs(folder_path, exist_ok=True)

    pict = f"{folder_path}/{current_time}.jpg"
    cv2.imwrite(pict, frame)

    img = Image(pict)
    img.height = 300
    img.width = 400
    
    row += 20
    ws[f"{col_time}{row}"] = current_time
    ws[f"{col_type}{row}"] = type
    ws.add_image(img, f"{col_image}{row}")
    wb.save(f"result_{start_time}.xlsx")
    generate_pdf(type, current_time, pict)
    Thread(target=send_telegram_message, args=(f"Terdeteksi kejadian {type.capitalize()} pada {current_time}. Mohon segera diperiksa kondisi terkininya",)).start()
    Thread(target=send_telegram_photo, args=(pict,)).start()


Thread(target=listen_to_bot, daemon=True).start()
>>>>>>> 6708c07 (integrate w software)

def filter_keypoints(keypoints):
    if np.all(keypoints == 0):
        return False
    if np.any(
        keypoints[
            [
                COCO_KEYPOINTS_NO_CONF.index(k)
                for k in [
                    "left_shoulder_x",
                    "left_shoulder_y",
                    "right_shoulder_x",
                    "right_shoulder_y",
                    "left_hip_x",
                    "left_hip_y",
                    "right_hip_x",
                    "right_hip_y",
                    "left_knee_x",
                    "left_knee_y",
                    "right_knee_x",
                    "right_knee_y",
                    "left_ankle_x",
                    "left_ankle_y",
                    "right_ankle_x",
                    "right_ankle_y",
                ]
            ]
        ]
        == 0
    ):
        return False
    return True
<<<<<<< HEAD


=======
>>>>>>> 6708c07 (integrate w software)
def process_frame(results, tracker_id, lstm_model, scaler):
    global rolling_buffer

    if tracker_id not in rolling_buffer:
        rolling_buffer[tracker_id] = []  # Initialize buffer for a new person

    keypoints = extract_keypoints([results])
    if keypoints:
        for kp in keypoints:
            if kp.person_index == tracker_id:
                keypoint_coord = []
                for keypoint_name in COCO_KEYPOINTS_NO_CONF:
                    value = getattr(kp, keypoint_name, None)
                    keypoint_coord.append(value)

                if filter_keypoints(np.array(keypoint_coord)):
                    rolling_buffer[tracker_id].append(keypoint_coord)
                    rolling_buffer[tracker_id] = rolling_buffer[tracker_id][
                        -MAX_TIMESTEPS:
                    ]

    print(f"Tracker ID: {tracker_id}, Frames: {len(rolling_buffer[tracker_id])}")
    # Predict only if we have at least `min_frames` in the buffer
    if len(rolling_buffer[tracker_id]) >= MIN_FRAMES:
        frame_data = np.array(rolling_buffer[tracker_id])
        frame_data = frame_data.reshape(-1, len(COCO_KEYPOINTS_NO_CONF))
        frame_data_df = pd.DataFrame(frame_data, columns=COCO_KEYPOINTS_NO_CONF)
        frame_data = scaler.transform(frame_data_df)

        # Pad or truncate to max_timesteps as required
        frame_data = pad_sequences(
            [frame_data],
            maxlen=MAX_TIMESTEPS,
            dtype="float32",
            padding="post",
            truncating="post",
        )

        # Predict the class
        prediction = lstm_model.predict(frame_data)
        predicted_class = np.argmax(prediction, axis=1)[0]
        return labels[predicted_class]

    return "Unknown"  # No prediction if insufficient frames

<<<<<<< HEAD

def process_video(
    video_path, yolo_model, lstm_model, scaler, show_output=True, save_output=False
):
    cap = cv2.VideoCapture(video_path)
=======
def lstm_results_loop(keypoint_results, person_labels, frame):
    for result in keypoint_results:
        for obj in result.boxes:
            tracker_id = obj.id
            if tracker_id is not None:
                tracker_id = tracker_id.item()
            else:
                continue

            label = process_frame(result, tracker_id, lstm_model, scaler)
            if label:
                person_labels[tracker_id] = label

                # Draw bounding box and label
                bbox = obj.xyxy[0].numpy().astype(int)
                cv2.rectangle(
                    frame, (bbox[0], bbox[1]), (bbox[2], bbox[3]), (0, 255, 0), 2
                )
                cv2.putText(
                    frame,
                    label,
                    (bbox[0], bbox[3] - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.9,
                    (0, 255, 0),
                    2,
                )
    return frame, person_labels

async def process_video(
    capture_device, websocket, show_output=True, save_output=False
):
    cap = capture_device
>>>>>>> 6708c07 (integrate w software)
    person_labels = {}

    frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = cap.get(cv2.CAP_PROP_FPS)
    if save_output:
        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        output_folder = os.path.join(os.getcwd(), "lstm_output")
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        output_path = os.path.join(
<<<<<<< HEAD
            output_folder, os.path.basename(video_path).replace(".mp4", "_output.mp4")
        )
        out = cv2.VideoWriter(output_path, fourcc, fps, (frame_width, frame_height))

    while cap.isOpened():
        success, frame = cap.read()
        if not success:
            break

        frame = cv2.resize(frame, target_size)

        results = yolo_model.track(frame, persist=True)
        for result in results:
            for obj in result.boxes:
                tracker_id = obj.id
                if tracker_id is not None:
                    tracker_id = tracker_id.item()
                else:
                    continue

                label = process_frame(result, tracker_id, lstm_model, scaler)
                if label:
                    person_labels[tracker_id] = label

                    # Draw bounding box and label
                    bbox = obj.xyxy[0].numpy().astype(int)
                    cv2.rectangle(
                        frame, (bbox[0], bbox[1]), (bbox[2], bbox[3]), (0, 255, 0), 2
                    )
                    cv2.putText(
                        frame,
                        label,
                        (bbox[0], bbox[3] - 10),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.9,
                        (0, 255, 0),
                        2,
                    )

        if show_output and isinstance(frame, np.ndarray):
            cv2.imshow("YOLO11 Tracking", frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

        if save_output and isinstance(frame, np.ndarray):
            out.write(frame)

    cap.release()
    if show_output:
        cv2.destroyAllWindows()


if __name__ == "__main__":
    video_folder = os.path.join(base_folder, "Walking")
    video_files = sorted([f for f in os.listdir(video_folder) if f.endswith(".mp4")])

    for video_file in video_files:
        video_path = os.path.join(video_folder, video_file)
        print(f"Processing video: {video_file}")
        process_video(
            video_path,
            yolo_model,
            lstm_model,
            scaler,
            show_output=True,
            save_output=True,
        )
        print("Not enough frames for prediction.")

        # Pause between videos
        input("Press Enter to continue to the next video...")
=======
            output_folder, os.path.basename('output.mp4')
        )
        out = cv2.VideoWriter(output_path, fourcc, fps, (frame_width, frame_height))

    row = 1
    prev_frame_time = time.time()
    fall_frame = 0
    frame_count = 0
    
    while cap.isOpened():
        now = datetime.now()
        current_time = now.strftime("%H-%M-%S %d-%m-%Y")
        ret, frame = cap.read()

        if not ret:
            print("Video frame is empty or video processing has been successfully completed.")
            break

        # FPS counter
        new_frame_time = time.time()
        fps = 1 / (new_frame_time - prev_frame_time)
        prev_frame_time = new_frame_time

        # Tampilkan status deteksi pada frame
        status_text = "Deteksi: ON" if status_detection else "Deteksi: OFF"
        cv2.putText(frame, status_text, (20, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 1)
        cv2.putText(frame, f"FPS: {int(fps)}", (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 1)

        if status_detection == True:
            fall_results = fall_model(frame, conf=0.8) 
            keypoint_results = keypoint_model.track(frame, persist=True)
            frame_count += 1

            for result in fall_results:
                frame = result.plot()
                print("frame_count: ", frame_count)
                class_id_list = list(result.boxes.cls)
                if frame_count > 0 and 0 in class_id_list:
                    fall_frame += 1
                    print("fall_frame: ", fall_frame)
                    if fall_frame == 20:
                        print("=======================")
                        print("=====Fall detected=====")
                        print("=======================")
                        process_telegram(frame, 'fall', start_time, current_time)
                        await websocket.send_text("fall")
                        fall_frame = 0
                        frame_count = -1000
                else:
                    fall_frame = 0
                    
            frame, person_labels = lstm_results_loop(keypoint_results, person_labels, frame)
            
            # person_labels = { [id : label], [id2 : label2] }
            for state in person_labels.values():
                if state != "Walking" or state != "Unknown":
                    print(f"Detected state: {state}")
                    process_telegram(frame, state.lower(), start_time, current_time, row)
                    await websocket.send_text(state.lower())

        if show_output and isinstance(frame, np.ndarray):
            cv2.imshow("YOLO11 Tracking", frame)
        if save_output and isinstance(frame, np.ndarray):
            out.write(frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break
    
async def app(scope, receive, send):
    websocket = WebSocket(scope=scope, receive=receive, send=send)
    await websocket.accept()
    await asyncio.sleep(5)
    
    try:
        process_video(
            cap,
            websocket,
            show_output=True,
            save_output=False,
        )
    except WebSocketDisconnect:
        print("WebSocket connection was disconnected.")
    except Exception as e:
        print(f"Error during detection: {e}")
    finally:
        cap.release()
        cv2.destroyAllWindows()
        await websocket.close()
>>>>>>> 6708c07 (integrate w software)
